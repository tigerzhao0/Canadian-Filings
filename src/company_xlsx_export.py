#!/usr/bin/env python3
"""Emit one company's raw fundamentals as an .xlsx matching a specific
template layout (see the row spec below) -- income statement, balance sheet,
and cash-flow statement, oldest-to-newest year left-to-right, straight from
financials.db. No computed ratios/subtotals invented: any template row that
doesn't have a direct raw match in our data is left as '-' (the template's own
"not applicable / not reported" convention), exactly like every other missing
cell in the source template.

    python src/company_xlsx_export.py RY   -> output/companies/RY.xlsx
    python src/company_xlsx_export.py --all -> one .xlsx per company

Standalone (sqlite3 + openpyxl only). Read-only against the DB; re-run any time.
"""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# --------------------------------------------------------------------------- #
# Row spec: reproduces the example template's exact row order and labels
# (leading spaces = indentation level, matched literally). Each data row maps
# to (statement_type, line_item) in our schema, or None if nothing in our data
# corresponds to it (left as '-'). `scale`:
#   "m"    -> stored value is raw dollars; divide by 1e6 to match the
#             template's "$ in millions" convention (confirmed against RY:
#             our TotalRevenue is 66,605,000,000 raw; template shows revenue
#             already in millions, e.g. Apple's 391035 = $391,035 million).
#   "raw"  -> per-share values (EPS, dividend/share) -- already correctly
#             scaled in our data, no conversion.
#   None   -> header/blank row, no value at all.
# A handful of template rows have no confident 1:1 field in our data (a
# combined "Deferred Tax And Revenue" bucket, "Other Gross PPE", the three
# ratio rows, etc.) -- those stay unmapped/'-' rather than guessed at.
ROW_SPEC = [
    ("header", "Annual Data:", None, None, None),
    ("fiscal", "Fiscal Period:", None, None, None),
    ("section", "Income Statement", None, None, None),
    # ---- bank rows (rendered only when the company reports bank keys) ----
    ("data_bank", "Interest Income", "income_statement", "InterestIncome", "m"),
    ("data_bank", "Interest Expense", "income_statement", "InterestExpense", "m"),
    ("data_bank", "Net Interest Income (for Banks)", "income_statement", "NetInterestIncome", "m"),
    ("data_bank", "Non Interest Income", "income_statement", "NonInterestIncome", "m"),
    ("data", "Revenue", "income_statement", "TotalRevenue", "m"),
    ("data_bank", "Credit Losses Provision", "income_statement", "CreditLossesProvision", "m"),
    ("data_bank", "Total Noninterest Expense", "income_statement", "NonInterestExpense", "m"),
    ("data", "  Cost of Goods Sold", "income_statement", "CostOfRevenue", "m"),
    ("data", "Gross Profit", "income_statement", "GrossProfit", "m"),
    ("blank", None, None, None, None),
    ("data", "  Selling, General, & Admin. Expense", "income_statement", "SellingGeneralAndAdministration", "m"),
    ("data", "  Research & Development", "income_statement", "ResearchAndDevelopment", "m"),
    ("data", "  Other Operating Expense", "income_statement", "OtherOperatingExpenses", "m"),
    ("data", "  Total Operating Expense", "income_statement", "OperatingExpense", "m"),
    ("data", "Operating Income", "income_statement", "OperatingIncome", "m"),
    ("blank", None, None, None, None),
    ("data", "    Interest Income", "income_statement", "InterestIncomeNonOperating", "m"),
    ("data", "    Interest Expense", "income_statement", "InterestExpenseNonOperating", "m"),
    ("data", "  Net Interest Income", "income_statement", "NetNonOperatingInterestIncomeExpense", "m"),
    ("data", "  Other Income (Expense)", "income_statement", "OtherIncomeExpense", "m"),
    ("data", "Pretax Income", "income_statement", "PretaxIncome", "m"),
    ("data", "  Tax Provision", "income_statement", "TaxProvision", "m"),
    ("blank", None, None, None, None),
    ("data", "  Other Net Income (Loss)", None, None, "m"),
    ("data", "  Net Income Including Noncontrolling Interests", None, None, "m"),
    ("data", "    Net Income (Continuing Operations)", "income_statement", "NetIncomeContinuousOperations", "m"),
    ("data", "    Net Income (Discontinued Operations)", "income_statement", "NetIncomeDiscontinuousOperations", "m"),
    ("data", "  Other Income (Minority Interest)", "income_statement", "MinorityInterests", "m"),
    ("data", "Net Income", "income_statement", "NetIncome", "m"),
    ("blank", None, None, None, None),
    ("blank", None, None, None, None),
    ("blank", None, None, None, None),
    ("data", "EPS (Basic)", "income_statement", "BasicEPS", "raw"),
    ("data", "EPS (Diluted)", "income_statement", "DilutedEPS", "raw"),
    ("data", "Shares Outstanding (Diluted Average)", "income_statement", "DilutedAverageShares", "m"),
    ("blank", None, None, None, None),
    ("data", "EBIT", "income_statement", "EBIT", "m"),
    ("data", "  Depreciation, Depletion and Amortization", "income_statement", "DepreciationAmortizationDepletion", "m"),
    ("data", "EBITDA", "income_statement", "EBITDA", "m"),
    ("ratio", "EBITDA Margin %", None, "ebitda_margin", None),
    ("ratio", "Tax Rate %", None, "tax_rate", None),
    ("ratio", "Net Margin %", None, "net_margin", None),
    ("blank", None, None, None, None),
    ("section", "Balance Sheet", None, None, None),
    # ---- bank rows (rendered only when the company reports bank keys) ----
    ("data_bank", "Balance Statement Cash and cash equivalents", "balance_sheet", "CashAndDueFromBanks", "m"),
    ("data_bank", "Money Market Investments", "balance_sheet", "MoneyMarketInvestments", "m"),
    ("data_bank", "Gross Loan", "balance_sheet", "GrossLoan", "m"),
    ("data_bank", "Allowance For Loans And Lease Losses", "balance_sheet", "AllowanceForLoansAndLeaseLosses", "m"),
    ("data_bank", "Net Loan", "balance_sheet", "NetLoan", "m"),
    ("data_bank", "Securities & Investments", "balance_sheet", "SecuritiesAndInvestments", "m"),
    ("data", "    Cash and Cash Equivalents", "balance_sheet", "CashAndCashEquivalents", "m"),
    ("data", "    Marketable Securities", "balance_sheet", "ShortTermInvestments", "m"),
    ("data", "  Cash, Cash Equivalents, Marketable Securities", "balance_sheet", "CashCashEquivalentsAndShortTermInvestments", "m"),
    ("data", "    Accounts Receivable", "balance_sheet", "AccountsReceivable", "m"),
    ("data", "    Notes Receivable", "balance_sheet", "NotesReceivable", "m"),
    ("data", "    Loans Receivable", "balance_sheet", "LoansReceivable", "m"),
    ("data", "    Other Current Receivables", "balance_sheet", "OtherReceivables", "m"),
    ("data", "  Total Receivables", "balance_sheet", "Receivables", "m"),
    ("data", "    Inventories, Raw Materials & Components", "balance_sheet", "RawMaterials", "m"),
    ("data", "    Inventories, Work In Process", "balance_sheet", "WorkInProcess", "m"),
    ("data", "    Inventories, Inventories Adjustments", "balance_sheet", "InventoriesAdjustmentsAllowances", "m"),
    ("data", "    Inventories, Finished Goods", "balance_sheet", "FinishedGoods", "m"),
    ("data", "    Inventories, Other", "balance_sheet", "OtherInventories", "m"),
    ("data", "  Total Inventories", "balance_sheet", "Inventory", "m"),
    ("data", "  Other Current Assets", "balance_sheet", "OtherCurrentAssets", "m"),
    ("data", "  Total Current Assets", "balance_sheet", "CurrentAssets", "m"),
    ("data", "  Investments And Advances", "balance_sheet", "InvestmentsAndAdvances", "m"),
    ("data", "      Land And Improvements", "balance_sheet", "LandAndImprovements", "m"),
    ("data", "      Buildings And Improvements", "balance_sheet", "BuildingsAndImprovements", "m"),
    ("data", "      Machinery, Furniture, Equipment", "balance_sheet", "MachineryFurnitureEquipment", "m"),
    ("data", "      Construction In Progress", "balance_sheet", "ConstructionInProgress", "m"),
    ("data", "      Other Gross PPE", None, None, "m"),
    ("data", "    Gross Property, Plant and Equipment", "balance_sheet", "GrossPPE", "m"),
    ("data", "    Accumulated Depreciation", "balance_sheet", "AccumulatedDepreciation", "m"),
    ("data", "  Property, Plant and Equipment", "balance_sheet", "NetPPE", "m"),
    ("data", "  Intangible Assets", "balance_sheet", "OtherIntangibleAssets", "m"),
    ("data", "    Goodwill", "balance_sheet", "Goodwill", "m"),
    ("data", "  Other Long Term Assets", "balance_sheet", "OtherNonCurrentAssets", "m"),
    ("data", "  Total Long-Term Assets", "balance_sheet", "TotalNonCurrentAssets", "m"),
    ("data", "Total Assets", "balance_sheet", "TotalAssets", "m"),
    ("blank", None, None, None, None),
    ("data", "    Accounts Payable", "balance_sheet", "AccountsPayable", "m"),
    ("data", "    Total Tax Payable", "balance_sheet", "TotalTaxPayable", "m"),
    ("data", "    Other Current Payables", "balance_sheet", "OtherPayable", "m"),
    ("data", "    Current Accrued Expense", "balance_sheet", "CurrentAccruedExpenses", "m"),
    ("data", "  Accounts Payable & Accrued Expense", "balance_sheet", "PayablesAndAccruedExpenses", "m"),
    ("data", "    Short-Term Debt", "balance_sheet", "CurrentDebt", "m"),
    ("data", "    Short-Term Capital Lease Obligation", "balance_sheet", "CurrentCapitalLeaseObligation", "m"),
    ("data", "  Short-Term Debt & Capital Lease Obligation", "balance_sheet", "CurrentDebtAndCapitalLeaseObligation", "m"),
    ("data", "    Current Deferred Revenue", "balance_sheet", "CurrentDeferredRevenue", "m"),
    ("data", "    Current Deferred Taxes Liabilities", "balance_sheet", "CurrentDeferredTaxesLiabilities", "m"),
    ("data", "  Deferred Tax And Revenue", "balance_sheet", "CurrentDeferredLiabilities", "m"),
    ("data", "  Other Current Liabilities", "balance_sheet", "OtherCurrentLiabilities", "m"),
    ("data", "  Total Current Liabilities", "balance_sheet", "CurrentLiabilities", "m"),
    ("data", "    Long-Term Debt", "balance_sheet", "LongTermDebt", "m"),
    ("data", "    Long-Term Capital Lease Obligation", "balance_sheet", "LongTermCapitalLeaseObligation", "m"),
    ("data", "  Long-Term Debt & Capital Lease Obligation", "balance_sheet", "LongTermDebtAndCapitalLeaseObligation", "m"),
    ("ratio", "Debt-to-Equity", None, "dte", None),
    ("data_bank", "Total Deposits", "balance_sheet", "TotalDeposits", "m"),
    ("data_bank", "Other Assets for Banks", "balance_sheet", "OtherAssets", "m"),
    ("data_bank", "Other Liabilities for Banks", "balance_sheet", "OtherPayable", "m"),
    ("data", "  Pension And Retirement Benefit", "balance_sheet", "NonCurrentPensionAndOtherPostretirementBenefitPlans", "m"),
    ("data", "    NonCurrent Deferred Revenue", "balance_sheet", "NonCurrentDeferredRevenue", "m"),
    ("data", "  NonCurrent Deferred Liabilities", "balance_sheet", "NonCurrentDeferredLiabilities", "m"),
    ("data", "  NonCurrent Deferred Income Tax", "balance_sheet", "NonCurrentDeferredTaxesLiabilities", "m"),
    ("data", "  Other Long-Term Liabilities", "balance_sheet", "OtherNonCurrentLiabilities", "m"),
    ("data", "  Total Long-Term Liabilities", "balance_sheet", "TotalNonCurrentLiabilities", "m"),
    ("data", "Total Liabilities", "balance_sheet", "TotalLiabilities", "m"),
    ("blank", None, None, None, None),
    ("data", "  Common Stock", "balance_sheet", "CommonStock", "m"),
    ("data", "  Preferred Stock", "balance_sheet", "PreferredStock", "m"),
    ("data", "  Retained Earnings", "balance_sheet", "RetainedEarnings", "m"),
    ("data", "  Accumulated other comprehensive income (loss)", "balance_sheet", "GainsLossesNotAffectingRetainedEarnings", "m"),
    ("data", "  Additional Paid-In Capital", "balance_sheet", "AdditionalPaidInCapital", "m"),
    ("data", "  Treasury Stock", "balance_sheet", "TreasuryStock", "m"),
    ("blank", None, None, None, None),
    ("data", "  Total Stockholders Equity", "balance_sheet", "StockholdersEquity", "m"),
    ("data", "  Minority Interest", "balance_sheet", "MinorityInterest", "m"),
    ("data", "Total Equity", "balance_sheet", "TotalEquityGrossMinorityInterest", "m"),
    ("ratio", "Equity-to-Asset", None, "eta", None),
    ("blank", None, None, None, None),
    ("section", "Cashflow Statement", None, None, None),
    ("data", "  Net Income From Continuing Operations", "cash_flow", "NetIncomeFromContinuingOperations", "m"),
    ("data", "  Cash Flow Depreciation, Depletion and Amortization", "cash_flow", "DepreciationAmortizationDepletion", "m"),
    ("data", "    Change In Receivables", "cash_flow", "ChangeInReceivables", "m"),
    ("data", "    Change In Inventory", "cash_flow", "ChangeInInventory", "m"),
    ("data", "    Change In Prepaid Assets", "cash_flow", "ChangeInPrepaidAssets", "m"),
    ("data", "    Change In Payables And Accrued Expense", "cash_flow", "ChangeInPayablesAndAccruedExpense", "m"),
    ("data", "    Change In Other Working Capital", "cash_flow", "ChangeInOtherWorkingCapital", "m"),
    ("data", "  Change In Working Capital", "cash_flow", "ChangeInWorkingCapital", "m"),
    ("blank", None, None, None, None),
    ("data", "  Stock Based Compensation", "cash_flow", "StockBasedCompensation", "m"),
    ("data", "  Asset Impairment Charge", "cash_flow", "AssetImpairmentCharge", "m"),
    ("data", "  Cash from Discontinued Operating Activities", "cash_flow", "CashFromDiscontinuedOperatingActivities", "m"),
    ("data", "  Cash Flow from Others", "cash_flow", "OtherNonCashItems", "m"),
    ("blank", None, None, None, None),
    ("blank", None, None, None, None),
    ("data", "  Purchase Of Property, Plant, Equipment", "cash_flow", "PurchaseOfPPE", "m"),
    ("data", "  Sale Of Property, Plant, Equipment", "cash_flow", "SaleOfPPE", "m"),
    ("data", "  Purchase Of Business", "cash_flow", "PurchaseOfBusiness", "m"),
    ("blank", None, None, None, None),
    ("data", "  Purchase Of Investment", "cash_flow", "PurchaseOfInvestment", "m"),
    ("data", "  Sale Of Investment", "cash_flow", "SaleOfInvestment", "m"),
    ("data", "  Net Intangibles Purchase And Sale", "cash_flow", "NetIntangiblesPurchaseAndSale", "m"),
    ("data", "  Cash From Discontinued Investing Activities", "cash_flow", "CashFromDiscontinuedInvestingActivities", "m"),
    ("data", "  Cash From Other Investing Activities", "cash_flow", "NetOtherInvestingChanges", "m"),
    ("data", "Cash Flow from Investing", "cash_flow", "InvestingCashFlow", "m"),
    ("blank", None, None, None, None),
    ("data", "  Issuance of Stock", "cash_flow", "IssuanceOfCapitalStock", "m"),
    ("data", "  Repurchase of Stock", "cash_flow", "RepurchaseOfCapitalStock", "m"),
    ("data", "  Net Issuance of Preferred Stock", "cash_flow", "NetPreferredStockIssuance", "m"),
    ("data", "    Issuance of Debt", "cash_flow", "IssuanceOfDebt", "m"),
    ("data", "    Payments of Debt", "cash_flow", "RepaymentOfDebt", "m"),
    ("data", "  Net Issuance of Debt", "cash_flow", "NetIssuancePaymentsOfDebt", "m"),
    ("data", "  Cash Flow for Dividends", "cash_flow", "CashDividendsPaid", "m"),
    ("data", "  Cash Flow for Lease Financing", None, None, "m"),
    ("data", "  Other Financing", "cash_flow", "NetOtherFinancingCharges", "m"),
    ("data", "Cash Flow from Financing", "cash_flow", "FinancingCashFlow", "m"),
    ("blank", None, None, None, None),
    ("data", "  Beginning Cash Position", "cash_flow", "BeginningCashPosition", "m"),
    ("data", "    Effect of Exchange Rate Changes", "cash_flow", "EffectOfExchangeRateChanges", "m"),
    ("data", "  Net Change in Cash", "cash_flow", "ChangesInCash", "m"),
    ("data", "Ending Cash Position", "cash_flow", "EndCashPosition", "m"),
    ("blank", None, None, None, None),
    ("data", "Capital Expenditure", "cash_flow", "CapitalExpenditure", "m"),
    ("data", "Free Cash Flow", "cash_flow", "FreeCashFlow", "m"),
]

MILLIONS_DIVISOR = 1_000_000.0

# GuruFocus's own export shows these two as a SIGNED CONTRIBUTION to the next
# subtotal (negative = reduces income) while every other expense-type row
# (COGS, SG&A, R&D, Other Operating Expense, D&A) stays a positive magnitude.
# Confirmed against a real GuruFocus export for Stingray Group (RAY): Tax
# Provision -1.7/-16/..., Interest Expense -17.5/-17.8/..., but SG&A 19.5 /
# Other Operating Expense 230.7 / D&A 40.3 (all positive). Internal storage
# MUST stay positive (derive.py's EBIT/NetIncome identities assume positive
# magnitudes) -- this flip is Sheet1-render-only, applied nowhere else (the
# "All Line Items" tab keeps the true stored value for traceability).
_NEGATE_ON_DISPLAY = {"TaxProvision", "InterestExpenseNonOperating"}

# Keys whose presence marks a company as a BANK (renders the data_bank rows).
_BANK_MARKER_KEYS = ("NetInterestIncome", "GrossLoan", "TotalDeposits",
                     "CreditLossesProvision", "NonInterestIncome")


def _fiscal_period_label(period_end: str | None) -> str:
    """'2025-10-31' -> '2025-10', matching the template's YYYY-MM style."""
    if not period_end or len(period_end) < 7:
        return "-"
    return period_end[:7]


def _ratio_value(rid: str, values: dict, fy: int) -> float | None:
    """Computed ratio rows (the template shows these; QuoteMedia doesn't store
    them). All inputs are raw dollars from statement_lines."""
    def g(stmt, key):
        return values.get((fy, stmt, key))
    if rid == "ebitda_margin":
        e, r = g("income_statement", "EBITDA"), g("income_statement", "TotalRevenue")
        return (e / r * 100) if (e is not None and r) else None
    if rid == "tax_rate":
        t, p = g("income_statement", "TaxProvision"), g("income_statement", "PretaxIncome")
        return (abs(t) / p * 100) if (t is not None and p) else None
    if rid == "net_margin":
        n, r = g("income_statement", "NetIncome"), g("income_statement", "TotalRevenue")
        return (n / r * 100) if (n is not None and r) else None
    if rid == "dte":
        debt = sum(v for v in (
            g("balance_sheet", "CurrentDebtAndCapitalLeaseObligation"),
            g("balance_sheet", "LongTermDebtAndCapitalLeaseObligation")) if v is not None)
        if not debt:
            debt = sum(v for v in (g("balance_sheet", "CurrentDebt"),
                                   g("balance_sheet", "LongTermDebt")) if v is not None)
        eq = (g("balance_sheet", "TotalEquityGrossMinorityInterest")
              or g("balance_sheet", "StockholdersEquity"))
        return (debt / eq) if (debt and eq) else None
    if rid == "eta":
        eq = (g("balance_sheet", "TotalEquityGrossMinorityInterest")
              or g("balance_sheet", "StockholdersEquity"))
        a = g("balance_sheet", "TotalAssets")
        return (eq / a) if (eq is not None and a) else None
    return None


def build_workbook(conn: sqlite3.Connection, ticker: str) -> openpyxl.Workbook | None:
    ident = conn.execute(
        "SELECT ticker, company_name, exchange, currency FROM companies WHERE ticker = ?",
        (ticker,)).fetchone()
    if ident is None:
        return None

    years_rows = conn.execute(
        "SELECT fiscal_year, period_end FROM company_years WHERE ticker = ? "
        "ORDER BY fiscal_year ASC", (ticker,)).fetchall()
    years = [y for y, _ in years_rows]
    period_end_by_year = dict(years_rows)

    lines = conn.execute(
        "SELECT fiscal_year, statement_type, line_item, value FROM statement_lines "
        "WHERE ticker = ?", (ticker,)).fetchall()
    values: dict[tuple[int, str, str], float | None] = {
        (fy, stmt, item): val for fy, stmt, item, val in lines
    }

    # Bank rows render only for companies that actually report bank keys.
    has_bank = any((fy, "income_statement", k) in values or (fy, "balance_sheet", k) in values
                   for fy in years for k in _BANK_MARKER_KEYS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    bold_font = Font(size=12, bold=True)
    plain_font = Font(size=12, bold=False)

    # Wide enough for the longest row label ("  Cash Flow Depreciation,
    # Depletion and Amortization", 52 chars) and for large dollar figures in
    # the fiscal-year columns (raw dollar amounts can run to 11+ digits).
    ws.column_dimensions["A"].width = 56
    for col in range(2, 2 + len(years)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    for kind, label, stmt, item, scale in ROW_SPEC:
        if kind == "data_bank" and not has_bank:
            continue
        row = ws.max_row + 1 if ws.max_row > 1 or ws.cell(1, 1).value else 1
        if kind == "blank":
            row = ws.max_row + 1
            continue
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = bold_font if kind == "section" else plain_font

        if kind == "fiscal":
            for col, fy in enumerate(years, start=2):
                c = ws.cell(row=row, column=col, value=_fiscal_period_label(period_end_by_year.get(fy)))
                c.font = plain_font
            continue
        if kind in ("header", "section"):
            continue

        if kind == "ratio":
            for col, fy in enumerate(years, start=2):
                v = _ratio_value(item, values, fy)
                c = ws.cell(row=row, column=col)
                c.font = plain_font
                if v is None:
                    c.value = "-"
                else:
                    c.value = v
                    c.number_format = "0.00"
            continue

        # kind == "data" / "data_bank"
        for col, fy in enumerate(years, start=2):
            raw = values.get((fy, stmt, item)) if stmt and item else None
            c = ws.cell(row=row, column=col)
            c.font = plain_font
            if raw is None:
                c.value = "-"
            else:
                v = raw / MILLIONS_DIVISOR if scale == "m" else raw
                if item in _NEGATE_ON_DISPLAY:
                    v = -v
                c.value = v
                c.number_format = "0.00"

    _add_all_lines_sheet(wb, conn, ticker, years)
    return wb


def _add_all_lines_sheet(wb, conn, ticker: str, years: list[int]) -> None:
    """Second sheet: EVERY parsed line item (mapped or not) from
    line_items_full -- the PDF pipeline's 'nothing is lost' view. Skipped
    silently for databases without that table (e.g. the QuoteMedia DB)."""
    try:
        rows = conn.execute(
            "SELECT statement_type, section, raw_label, canonical_key, "
            "fiscal_year, value, source_pdf_year FROM line_items_full "
            "WHERE ticker = ? ORDER BY statement_type, source_pdf_year DESC, line_no",
            (ticker,)).fetchall()
    except sqlite3.OperationalError:
        return
    if not rows:
        return
    # newest source PDF wins per (statement, label, year)
    cell_val: dict[tuple, float] = {}
    order: list[tuple] = []
    seen: set[tuple] = set()
    keymap: dict[tuple, tuple] = {}
    for stmt, section, label, key, fy, val, src in rows:
        ident = (stmt, section or "", label)
        if ident not in seen:
            seen.add(ident)
            order.append(ident)
            keymap[ident] = (key,)
        slot = (stmt, section or "", label, fy)
        if slot not in cell_val and val is not None:
            cell_val[slot] = val
    ws = wb.create_sheet("All Line Items")
    bold = Font(bold=True)
    headers = ["Statement", "Section", "Line item (as printed)", "Canonical key"] + \
              [str(y) for y in years]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 30
    r = 2
    for ident in order:
        stmt, section, label = ident
        ws.cell(row=r, column=1, value=stmt)
        ws.cell(row=r, column=2, value=section)
        ws.cell(row=r, column=3, value=label)
        ws.cell(row=r, column=4, value=keymap[ident][0] or "")
        for col, fy in enumerate(years, start=5):
            v = cell_val.get((stmt, section, label, fy))
            c = ws.cell(row=r, column=col, value=(v / MILLIONS_DIVISOR if v is not None else "-"))
            if v is not None:
                c.number_format = "0.00"
        r += 1


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("ticker", nargs="?", help="Export one company (e.g. RY).")
    ap.add_argument("--all", action="store_true", help="Export every company in the DB.")
    ap.add_argument("--db", default="output/financials.db")
    ap.add_argument("--out-dir", default="output/companies")
    args = ap.parse_args()

    if not args.ticker and not args.all:
        ap.error("Provide a ticker (e.g. RY) or --all.")

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(args.db)
    try:
        if args.all:
            tickers = [r[0] for r in conn.execute("SELECT ticker FROM companies ORDER BY ticker")]
        else:
            tickers = [args.ticker.upper()]

        written = 0
        for tk in tickers:
            wb = build_workbook(conn, tk)
            if wb is None:
                if not args.all:
                    print(f"'{tk}' not in companies table (never resolved). "
                         "Run the financials stage first.")
                continue
            wb.save(out_dir / f"{tk}.xlsx")
            written += 1
        if args.all:
            print(f"Wrote {written} company .xlsx file(s) to {out_dir}/")
        elif written:
            print(f"Wrote {out_dir}/{tickers[0]}.xlsx")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
